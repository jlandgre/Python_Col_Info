#Version 4/21/25
import os, sys
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import itertools

path_libs = os.getcwd() + os.sep + 'libs' + os.sep
if not path_libs in sys.path: sys.path.append(path_libs)
import pd_util

"""
================================================================================
ProjectTables Class -- this can be initialized as tbls to manage all data
tables for a project. __init__ instances a Table class for each table and
initializes lists column name mapping and for preflight checks

JDL 9/3/24
================================================================================
"""
class ProjectTables():
    """
    Collection of imported or generated data tables for a project
    JDL 9/26/24; Modified 4/9/25
    """
    def __init__(self, files, IsPrint=False):
        """
        Instance attributes including Table instances
        """
        self.IsPrint = IsPrint
        self.files = files

        #Instance project-specific tables and ColInfo
        self.InstanceTblObjs()
        self.InstanceAndImportColInfo()
        
    def InstanceTblObjs(self):
        """
        Instance project-specific tables 
        (example data tests/test_data/ folder)
        JDL 4/9/25
        """

        # Set import params for Excel files
        dImportParams = {'import_path':self.files.path_data,
                        'ftype': 'excel',
                        'sht':'data'}

        dImportParams['lst_files'] = 'Example1.xlsx'
        self.ExampleTbl1 = Table('ExampleTbl1', dImportParams)

        dImportParams['lst_files'] = 'Example2.xlsx'
        self.ExampleTbl2 = Table('ExampleTbl2', dImportParams)

        self.lstExcelImports = [self.ExampleTbl1, self.ExampleTbl2]

    def InstanceAndImportColInfo(self):
        """
        Instance and Import ColInfo table with metadata about variables
        JDL 4/9/25
        """
        # Instance the ColInfo table and import to tbls.ColInfo.df
        dImportParams = {'ftype':'excel',
                        'lst_files':'col_info.xlsx', 
                        'import_path':self.files.path_data,
                        'sht':'cols'}

        self.ColInfo = Table('ColInfo', dImportParams)
        self.ColInfo.ImportToTblDf()
        #self.ImportExcelInputs(lstExcelImports=[self.ColInfo])
        #self.ColInfo.ImportToTblDf_New()

    def ImportExcelInputs(self, lstExcelImports=None):
        """
        Read rows/cols input data - use pd_util.ImportExcel() to avoid importing 
        blank columns in sheet's Excel .UsedRange. Specify 
        tbl.dParseParams['col_last_df'] to specify where to truncate columns
        JDL 11/16/24 Add lstImports optional argument
        """
        if lstExcelImports is None: lstExcelImports = self.lstExcelImports

        for tbl in lstExcelImports:
            tbl.ImportExcelDf()

            if self.IsPrint:
                print('\nImported Excel', tbl.name, tbl.pf, tbl.sht)
                print(tbl.df)

    def ImportCSVInputs(self, lstImportsCSV=None):
        """
        Read rows/cols input data from CSV files (read directly to .df with no
        parsing of initially-imported .df_raw)
        JDL 3/31/25
        """
        if lstImportsCSV is None: lstImportsCSV = self.lstImportsCSV

        for tbl in lstImportsCSV:
            tbl.df = pd.read_csv(tbl.pf)

            if self.IsPrint:
                print('\nImported CSV', tbl.name, tbl.pf)
                print(tbl.df)

    def ImportRawInputs(self):
        """
        Read each table's raw data using openpyxl to work on sheets whose data 
        may not start at A1 (e.g. .df_raw requires parsing to .df)
        JDL 3/4/24; Modified 9/25/24
        """
        for tbl in self.lstRawImports:
            tbl.ImportExcelRaw()

            if self.IsPrint:
                print('\nImported Excel Raw', tbl.name, tbl.pf, tbl.sht)
                print(tbl.df)

class Table():
    """
    Attributes for a data table including import instructions and other
    metadata. Table instances are attributes of ProjectTables Class
    JDL Modified 4/8/25 refactor to fully use dImportParams and dParseParams
    """
    def __init__(self, name, dImportParams=None, dParseParams=None,):
                
        self.name = name #Table name

        # Dicts of import and parsing parameters
        self.dImportParams = dImportParams or {}
        self.dParseParams = dParseParams or {'parse_type':'none'}
        self.df_raw = pd.DataFrame() # temp raw data in .lst_dfs iteration
        self.df = pd.DataFrame()
        self.col_info = None

        # Temp variables for looping through files
        self.pf = None
        self.sht = None
        self.lst_dfs = None
        self.sht_type = None
        self.is_unstructured = None
        self.lst_dfs = None

    """
    ================================================================================
    ParseRawData Procedure
    If imported data are not rows/cols structured, parse .lst_dfs to .df using
    dParseParams inputs to guide parsing
    JDL 4/21/25
    ================================================================================
    """
    def ParseRawData(tbl):
        """
        Procedure to parse raw data for a given Table instance.
        """
        for df in tbl.lst_dfs:
            if tbl.dParseParams['parse_type'] == 'row_major':
                parse = RowMajorTbl(tbl, df)
                parse.ReadBlocksProcedure()
    """
    ================================================================================
    ApplyColInfo Procedure
    Apply col_info to tbl.df to rename and subset raw columns and set default index
    and data types
    JDL 4/21/25
    ================================================================================
    """
    # Parse df_raw to df using dParseParams
    """
    ================================================================================
    ImportToTblDf Procedure
    JDL 4/10/25 Rewritten to allow multisheet Excel and separate ingest/parse
    ================================================================================
    """
    def ImportToTblDf(self, lst_files=None):
        """
        Procedure to import file(s) + sheet(s) to self.df (structured rows/cols)
        or self.lst_dfs (unstructured) using .dImportParams and .dParseParams to
        set options

        Can directly specify lst_files as arg or as dImportParams['lst_files']
        Refactored JDL 4/10/25; Comments updated 4/21/25 to clarify
        """
        # Set lst_files based on dImportParams['lst_files'] or input arg    
        lst_files = self.SetLstFiles(lst_files)

        # Set file ingest parameters based on dImportParams or on default values
        self.SetFileIngestParams()

        # initialize list df's and temp df (temp if structured; or for parsing later)
        self.lst_dfs = []
        self.df_temp = pd.DataFrame()

        # Loop over input list of files to ingest
        for self.pf in lst_files:

            # Read from Excel single/multiple sheets self.pf; append to lst_dfs
            if self.dImportParams['ftype'] == 'excel':
                self.SetLstSheets()
                self.ReadExcelFileSheets()

            # Read from CSV self.pf; append to lst_dfs
            elif self.dImportParams['ftype'] == 'csv':
                self.ReadCSVFile()

            # Read from feather self.pf; append to lst_dfs
            elif self.dImportParams['ftype'] == 'feather':
                pass

        #Concat if rows/cols aka structured (e.g. no parsing needed)
        if not self.is_unstructured:
            self.df = pd.concat(self.lst_dfs, ignore_index=True)
            self.lst_dfs = []

    def SetLstFiles(self, lst_files):
        """
        Set lst_files based on input and dImportParams.
        """
        # If lst_files is not specified, use dImportParams['lst_files']
        if lst_files is None: lst_files = self.dImportParams['lst_files']

        # Convert to list if a single file is provided
        if not isinstance(lst_files, list): lst_files = [lst_files]

        # Optionally prepend import_path to each file name
        if 'import_path' in self.dImportParams:
            lst_files = [self.dImportParams['import_path'] + f for f in lst_files]
        return lst_files

    def SetFileIngestParams(self):
        """
        Set Table attributes for the current file 
        (concise vs referencing dict items and also factors in default vals if
        dict item not specified)
        JDL 4/10/25
        """
        self.is_unstructured = self.SetParseParam(False, 'is_unstructured')
        self.n_skip_rows = self.SetParseParam(0, 'n_skip_rows')
        self.parse_type = self.SetParseParam('none', 'parse_type')
        if self.dImportParams['ftype'] == 'excel':
            self.sht_type = self.SetImportParam('single', 'sht_type')

    def SetImportParam(self, valDefault, param_name):
        """
        Set default or non-default import parameter
        JDL 4/9/25
        """
        val = valDefault
        if param_name in self.dImportParams: val = self.dImportParams[param_name]
        return val

    def SetParseParam(self, valDefault, param_name):
        """
        Set default or non-default parsing parameter
        JDL 4/9/25
        """
        val = valDefault
        if param_name in self.dParseParams: val = self.dParseParams[param_name]
        return val

    def SetLstSheets(self):
        """
        Set .lst_sheets based on sht_type and sht in dImportParams
        (Called within iteration with self.pf file)
        JDL 4/10/25
        """
        self.lst_sheets = []

        if self.sht_type == 'single':

            # Set sheet name to either specified or 0 (e.g. first sheet)
            self.lst_sheets = [self.dImportParams.get('sht', 0)]

            # If sheet name is 0, reset it to first sheet name
            if self.lst_sheets[0] == 0:
                wb = load_workbook(filename=self.pf, read_only=True)
                self.lst_sheets[0] = wb.sheetnames[0]
                wb.close()

        elif self.sht_type == 'all':
            wb = load_workbook(filename=self.pf, read_only=True)
            self.lst_sheets = wb.sheetnames
            wb.close()
        
        elif self.sht_type == 'list':
            pass

        elif self.sht_type == 'regex':
            pass

        elif self.sht_type == 'startswith':
            pass

        elif self.sht_type == 'endswith':
            pass

        elif self.sht_type == 'contains':
            pass

    def ReadExcelFileSheets(self):
        """
        Loop through sheets in lst_sheets and read their data
        JDL 4/10/25
        """
        for self.sht in self.lst_sheets:
            self.ReadExcelSht()
            self.lst_dfs.append(self.df_temp)
            self.df_temp = pd.DataFrame()

    def ReadExcelSht(self):
        """
        Read data from the current sheet into a temporary DataFrame.
        """
        if self.is_unstructured:
            #self.df_temp = self.ImportExcelRaw()
            self.df_temp = pd.read_excel(self.pf, sheet_name=self.sht, header=None)

            # Negate Pandas inferring float data type for integers and NaNs for blanks
            if 'import_dtype' in self.dParseParams and self.dParseParams['import_dtype'] == str:
                self.df_temp = self.df_temp.astype(object)
                self.df_temp = self.df_temp.map(lambda x: None if pd.isna(x) \
                    else str(int(x)) if isinstance(x, float) and x.is_integer() else str(x))
        else:
            self.df_temp = pd.read_excel(self.pf, sheet_name=self.sht, skiprows=self.n_skip_rows)

    def ReadCSVFile(self):
        """
        Import current CSV file into a temporary df and append to lst_dfs
        JDL 4/10/25
        """
        if self.is_unstructured:
            # Read CSV without treating first row as headers
            self.df_temp = pd.read_csv(self.pf, header=None)
        else:
            # Read CSV with optional skiprows
            self.df_temp = pd.read_csv(self.pf, skiprows=self.n_skip_rows)

        # Append temp df to lst_dfs and re-initialize
        self.lst_dfs.append(self.df_temp)
        self.df_temp = pd.DataFrame()

    def ImportExcelDf_obsolete(self):
        """
        Import rows/cols homed table data from Excel to .df
        JDL 9/3/24; 11/16/24 to deal with non-specified .sht (import 1st sht)
        """
        # Use first sheet if self.sht is empty
        sht = self.sht if self.sht else 0  

        self.df = pd_util.dfExcelImport(self.pf, sht=sht, \
                                        IsDeleteBlankCols=True)
        
        #Optionally, drop columns after specified last column
        if self.dParseParams is not None and 'col_last_df' in self.dParseParams:
            col_last = self.dParseParams['col_last_df']
            try:
                idx_last = self.df.columns.get_loc(col_last)
                self.df = self.df.iloc[:, :idx_last+1]
            except KeyError:
                raise ValueError(f"Column {col_last} not found in", self.name)

    def ImportExcelRaw_obsolete(self):
        """
        Import unstructured data to .df_raw for parsing
        JDL Modified 9/26/24 to allow forcing str type for imported values
        """
        #Create workbook object and select sheet
        wb = load_workbook(filename=self.pf, read_only=True)
        sht = self.sht if self.sht != 0 else wb.sheetnames[0]
        ws = wb[sht]

        # Convert the data to a list and convert to a DataFrame
        data = ws.values
        df = pd.DataFrame(data)

        #Negate Pandas inferring float data type for integers and NaNs for blanks
        if 'import_dtype' in self.dParseParams and self.dParseParams['import_dtype'] == str:
            df = df.applymap(lambda x: None if pd.isna(x) \
                    else str(int(x)) if isinstance(x, float) and x.is_integer() \
                    else str(x))
        return df

    def ResetDefaultIndex(self, IsDrop=True):
        """
        Set or Reset df index to the default defined for the table
        JDL 2/20/24; Fix bug with else branch 9/3/24
        """
        if self.idx_col_name is None: return self.df
        if self.df.index.name is None:
            self.df = self.df.set_index(self.idx_col_name)
        else:
            self.df = self.df.reset_index(drop=IsDrop)
            self.df = self.df.set_index(self.idx_col_name)

class CheckInputs:
    """
    Check the tbls dataframes for errors
    (dummy initialization of preflight check)
    """
    def __init__(self, tbls, IsPrint=True):
        self.tbls = tbls
        self.IsPrint = IsPrint

        #preflight.CheckDataFrame Class --instanced as needed in methods below
        self.ckdf = None    

"""
================================================================================
InterleavedColBlocksTbl Class - Data in interleaved, repeating column blocks
================================================================================
"""
class InterleavedColBlocksTbl():
    """
    Table contains initial metadata columns (row major) followed by interleaved
    repeating blocks of columns containing row major data. Block ID variable is
    in Row 1 first column of each block. Repeating variable names in Row 2
    """
    def __init__(self, tbl):

        #Raw DataFrame and column list parsed from raw data
        self.df_raw = tbl.df_raw

        # Start index to allow for initial, blank/unused columns
        self.idx_start = 0
        if 'idx_start' in tbl.dParseParams:
            self.idx_start = tbl.dParseParams['idx_start']

        # List of metadata columns preceding interleaved blocks
        self.n_cols_metadata = tbl.dParseParams['n_cols_metadata']

        # n columns per block
        self.n_cols_block = tbl.dParseParams['n_cols_block']

        self.df = pd.DataFrame()

        # Iteration variables
        self.idx_col_block_cur = None
        self.block_name_cur = None
        self.idx_col_cur = None

    def ParseInterleavedBlocksProcedure(self):
        """
        Procedure to parse interleaved blocks of columns
        JDL 3/17/25
        """
        self.SetDfMetadata()
        self.DeleteTrailingRows()
        self.TransferAllBlocks()
    
    def SetDfMetadata(self):
        """
        Set .df_metadata as a subset of .df_raw 
        JDL 
        """
        # Subset the metadata columns starting at idx_start
        col_last = self.idx_start + self.n_cols_metadata
        self.df_metadata = self.df_raw.iloc[2:, self.idx_start:col_last]
        self.df_metadata = self.df_metadata.reset_index(drop=True)
        self.df_metadata.columns = self.df_raw.iloc[1, self.idx_start:col_last]
        self.df_metadata.columns.name = None

    def DeleteTrailingRows(self):
        """
        Delete trailing rows with blank metadata
        JDL 3/17/25
        """
        # Find the index of last non-null metadata row
        idx_last = self.df_metadata.apply(lambda row: \
            not row.isnull().all(), axis=1).cumsum().idxmax()

        # Delete trailing rows from .df_metadata and corresponding .df_raw rows
        self.df_metadata = self.df_metadata.iloc[:idx_last + 1]
        self.df_raw = self.df_raw.iloc[:idx_last + 3]

    def TransferAllBlocks(self):
        """
        Transfer all blocks of columns to .df
        JDL 3/17/25
        """
        # Initialize .idx_col_cur
        self.idx_col_cur = self.idx_start + self.n_cols_metadata

        # Iteratively read blocks until end of columns
        while self.idx_col_cur < len(self.df_raw.columns) and \
            not pd.isna(self.df_raw.loc[0, self.idx_col_cur]):
            self.ReadWriteBlock()
            self.idx_col_cur += 1

    def ReadWriteBlock(self):
        """
        Read and write a block of columns to .df
        JDL 3/17/25
        """
        # Read the block name
        self.idx_col_block_cur = self.idx_col_cur
        self.block_name_cur = self.df_raw.loc[0, self.idx_col_block_cur]

        # Iteratively call ReadWriteColData for each column in the block
        for i in range(self.n_cols_block):
            self.idx_col_cur = self.idx_col_block_cur + i
            self.ReadWriteColData()

    def ReadWriteColData(self):
        """
        Transfer one column's data to .df by reading from a column block
        JDL 3/17/25
        """
        # Append .df_metadata to .df and write block name to new rows
        self.df = pd.concat([self.df, self.df_metadata], ignore_index=True)
        n_rows_new = len(self.df_metadata)
        self.df.loc[self.df.index[-n_rows_new:], 'block_name'] = self.block_name_cur

        # Write the variable name (.df_raw row index 1)
        var_name = self.df_raw.iloc[1, self.idx_col_cur]
        self.df.loc[self.df.index[-n_rows_new:], 'var_name'] = var_name

        # write values (as array to avoid index conflict between values and .df
        values = self.df_raw.iloc[2:, self.idx_col_cur].reset_index(drop=True)
        self.df.loc[self.df.index[-n_rows_new:], 'values'] = values.values
"""
================================================================================
RowMajorTbl Class - for parsing row major raw data single block
================================================================================
"""
class RowMajorTbl():
    """
    Description and Parsing Row Major Table initially embedded in tbl.df
    (imported with tbls.ImportInputs() or .ImportRawInputs() methods
    JDL 3/4/24; Modified 9/26/24
    """
    def __init__(self, tbl, df=None):
        
        # If df (raw data) is specified from tbl.lst_dfs iteration, use it
        if df is not None:
            self.df_raw = df
        else:
            self.df_raw = tbl.df_raw

        #List of df indices for rows where flag_start_bound is found
        self.start_bound_indices = []

        #Raw DataFrame and column list parsed from raw data
        #self.df_raw = tbl.df_raw

        #Table whose df is to be populated by parsing
        self.tbl = tbl
        self.lst_block_ids = []

        #Start, header, end, first data row indices for current block in loop
        self.idx_start_current = None
        self.idx_header_row = None
        self.idx_end_bound = None
        self.idx_start_data = None

        #Current block's columns and parsed data
        self.cols_df_block = []
        self.df_block = pd.DataFrame()
    """
    ================================================================================
    """
    def ReadBlocksProcedure(self):
        """
        Procedure to iteratively parse row major blocks
        JDL 9/26/24
        """
        # Append blank row at end of .df_raw (to ensure find last <blank> flag)
        self.AddTrailingBlankRow()

        #Create list of row indices with start bound flag
        self.SetStartBoundIndices()

        #Iteratively read blocks 
        for i in self.start_bound_indices:
            self.idx_start_current = i
            self.ParseBlockProcedure()

        #Extract block_id values if specified (Note: self arg is RowMajorTbl instance)
        self.tbl.df, self.lst_block_ids = RowMajorBlockID(self).ExtractBlockIDs

        #set default index
        #self.SetDefaultIndex()
        self.tbl.df = self.tbl.df.reset_index(drop=True)

        #Optionally stack parsed data (if .dParams['is_stack_parsed_cols']
        #self.StackParsedCols()

    def AddTrailingBlankRow(self):
        """
        Add a trailing blank row to self.df_raw (to ensure last <blank> flag to
        terminate last block)
        JDL 9/26/24
        """
        blank_row = pd.Series([np.nan] * len(self.df_raw.columns), index=self.df_raw.columns)
        self.df_raw = pd.concat([self.df_raw, pd.DataFrame([blank_row])], ignore_index=True)

    def SetStartBoundIndices(self):
        """
        Populate list of row indices whereflag_start_bound is found
        JDL 9/25/24
        """
        flag= self.tbl.dParseParams['flag_start_bound']
        icol = self.tbl.dParseParams['icol_start_bound']

        fil = self.df_raw.iloc[:, icol] == flag
        self.start_bound_indices = self.df_raw[fil].index.tolist()

    def SetDefaultIndex(self):
        """
        Set the table's default index
        JDL 3/4/24
        """
        self.tbl.df = self.tbl.df.set_index(self.tbl.idx_col_name)
    
    def StackParsedCols(self):
        """
        Optionally stack parsed columns from row major blocks
        JDL 9/25/24; Modified 4/23 remove from procedure/move to Apply ColInfo
        """
        is_stack = self.tbl.dParseParams.get('is_stack_parsed_cols', False)

        #xxx
        self.tbl.df = self.tbl.df.set_index('Answer Choices')
        self.tbl.idx_col_name = 'Answer Choices'
        print('\n', self.tbl.df)

        if is_stack:
            #xxx
            self.tbl.df = self.tbl.df.stack().reset_index()
            self.tbl.df = self.tbl.df.stack()

            print('\n', self.tbl.df)

            #Respecify the index column name and set default index
            self.tbl.df.columns = [self.tbl.idx_col_name, 'Metric', 'Value']
            self.SetDefaultIndex()

    def ParseBlockProcedure(self):
        """
        Parse the table and set self.df resulting DataFrame
        JDL 9/25/24; Modified 4/22/25
        """
        self.FindFlagEndBound()
        self.ReadHeader()
        self.SubsetDataRows()
        #self.SubsetCols()
        #self.RenameCols()
        #self.SetColumnOrder()

        #Concatenate into tbl.df and re-initialize df_block
        self.tbl.df = pd.concat([self.tbl.df, self.df_block], axis=0)
        self.df_block = pd.DataFrame()

    def FindFlagEndBound(self):
        """
        Find index of flag_end_bound
        JDL 3/4/24; modified 9/26/24
        """
        flag = self.tbl.dParseParams['flag_end_bound']
        icol = self.tbl.dParseParams['icol_end_bound']
        ioffset = self.tbl.dParseParams['idata_rowoffset_from_flag']

        #Start the search at the first data row based on data offset from flag
        i = self.idx_start_current + ioffset

        # search for specifie flag string/<blank> below row i
        if flag == '<blank>':
            self.idx_end_bound = self.df_raw.iloc[i:, icol].isnull().idxmax()
        else:
            self.idx_end_bound = self.df_raw.iloc[i:, icol].eq(flag).idxmax()

    def ReadHeader(self):
        """
        Read header based on iheader_rowoffset_from_flag.
        JDL 3/4/24; modified 9/26/24
        """
        # Calculate the header row index
        iheader_offset = self.tbl.dParseParams['iheader_rowoffset_from_flag']
        self.idx_header_row =  self.idx_start_current + iheader_offset

        # Set the column names (drop columns with blank header)
        self.cols_df_block = self.df_raw.iloc[self.idx_header_row].values

    def SubsetDataRows(self):
        """
        Subset raw data rows based on flags and idata_rowoffset_from_flag
        JDL 3/4/24; Modified 4/23/25
        """
        # Calculate the start index for the data
        self.idx_start_data = self.idx_start_current + \
            self.tbl.dParseParams['idata_rowoffset_from_flag']

        # Create df with block's data rows
        self.df_block = self.df_raw.iloc[self.idx_start_data:self.idx_end_bound]

        # Added 4/23 to replace doing this in SubsetCols()
        self.df_block.columns = self.cols_df_block

        # Added 4/23 drop columns with null column name and all null values
        fil = ~self.df_block.columns.isnull() & self.df_block.notna().any()
        self.df_block = self.df_block.loc[:, fil]

    def SubsetCols(self):
        """
        Use tbl.import_col_map to subset columns based on header.
        JDL 9/25/24
        """
        self.df_block.columns = self.cols_df_block

        #Use import_col_map if specified
        if len(self.tbl.import_col_map) > 0:
            cols_keep = list(self.tbl.import_col_map.keys())
            self.df_block = self.df_block[cols_keep]
        
        #Drop columns with blank (e.g. NaN) header
        else:
            self.df_block = self.df_block.dropna(axis=1, how='all')

    def RenameCols(self):
        """
        Optionally use tbl.import_col_map to rename columns.
        JDL 3/4/24; Modified 9/24/24
        """
        if len(self.tbl.import_col_map) > 0:
            self.df_block.rename(columns=self.tbl.import_col_map, inplace=True)

    def SetColumnOrder(self):
        """
        Set column order based on tbl.col_order Series
        JDL 4/22/25
        """
        if self.tbl.col_order is not None:
            self.df_block = self.df_block[self.tbl.col_order]


"""
================================================================================
RowMajorBlockID Class - sub to RowMajorTbl for extracting block_id values
================================================================================
"""
class RowMajorBlockID:
    def __init__(self, parse_instance):
        self.tbl = parse_instance.tbl

        #Index of first data row in .df_raw
        self.idx_start_data = parse_instance.idx_start_data

        #Raw df that is being parsed (.df_raw set in parse_instance.__init__())
        self.df_raw = parse_instance.df_raw

        #Within loop value for a block ID
        self.block_id_value = None

        #List of all block_id names
        self.df_cols_initial = self.tbl.df.columns.tolist()
        self.block_id_names = []

    @property
    def ExtractBlockIDs(self):
        """
        Property returns updated DataFrame and list of names.
        JDL 9/27/24
        """
        self.ExtractBlockIDsProcedure()
        return self.tbl.df, self.block_id_names
    """
    ============================================================================
    """
    def ExtractBlockIDsProcedure(self):
        """
        Procedure to extract block ID values from df_raw based on current block's
        data row index and dict list of block_id tuples: (block_id_name, row_offset,
        col_index) where row_offset is offset from idx_start_data and col_index is 
        absolute column index where each block_id value is found.
        JDL 9/27/24
        """
        #Convert to list if specified as one-item tuple
        self.ConvertTupleToList()

        #Iterate through block_id tuples and add columns to tbl.df
        for tup_block_id in self.tbl.dParseParams.get('block_id_vars', []):
            self.SetBlockIDValue(tup_block_id)
            #self.ReorderColumns() #Don't do as part of parsing -- ApplyColInfo takes care of

    def ConvertTupleToList(self):
        """
        If only one block_id, it can be specified as tuple; otherwise it's
        a list of tuples.
        JDL 9/27/24
        """
        if 'block_id_vars' in self.tbl.dParseParams:

            #If necessary, convert tuple to one-item list
            if isinstance(self.tbl.dParseParams['block_id_vars'], tuple):
                self.tbl.dParseParams['block_id_vars'] = \
                    [self.tbl.dParseParams['block_id_vars']]
            
    def SetBlockIDValue(self, tup_block_id):
        """
        Set internal values based current block_id tuple
        JDL 9/27/24; Modified 4/22/25 for RowMajorTbl refactor
        """
        name, row_offset = tup_block_id[0], tup_block_id[1]
        idx_row, idx_col = self.idx_start_data + row_offset, tup_block_id[2]

        #Set the current value and add the name list
        value_block_id = self.df_raw.iloc[idx_row, idx_col]
        self.block_id_names.append(name)
        self.tbl.df[name] = value_block_id

    def ReorderColumns(self):
        """
        Reorder so that block_id columns are first
        9/27/24
        """
        self.tbl.df = self.tbl.df[self.block_id_names + self.df_cols_initial]

